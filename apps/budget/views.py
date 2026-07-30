from datetime import date, datetime
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.mixins import AdminRequiredMixin
from .forms import CategorieChargeForm, OperationBudgetForm
from .models import CategorieCharge, OperationBudget


def _operations_filtrees(request):
    """Filtre partagé par la liste et l'export « filtre actif »."""
    queryset = OperationBudget.objects.select_related(
        'categorie_charge', 'seance__client', 'abonnement__client', 'mouvement_stock__produit'
    ).all()
    type_operation = request.GET.get('type', 'tous')
    if type_operation in ('entree', 'sortie'):
        queryset = queryset.filter(type_operation=type_operation)
    categorie = request.GET.get('categorie', 'toutes')
    if categorie != 'toutes':
        queryset = queryset.filter(categorie=categorie)
    return queryset


def _dates_rapport(request):
    """Bornes (incluses) du rapport d'opérations — défaut : depuis le 1er du
    mois en cours jusqu'à aujourd'hui, modifiable via les paramètres GET
    `debut`/`fin` (format YYYY-MM-DD, celui des <input type="date">)."""
    aujourdhui = date.today()
    try:
        date_debut = datetime.strptime(request.GET.get('debut', ''), '%Y-%m-%d').date()
    except ValueError:
        date_debut = aujourdhui.replace(day=1)
    try:
        date_fin = datetime.strptime(request.GET.get('fin', ''), '%Y-%m-%d').date()
    except ValueError:
        date_fin = aujourdhui
    return date_debut, date_fin


def _operations_rapport(request):
    """Filtre partagé par l'écran de rapport et son export — mêmes filtres
    type/catégorie que le journal, restreints à la période choisie."""
    date_debut, date_fin = _dates_rapport(request)
    return _operations_filtrees(request).filter(date__date__gte=date_debut, date__date__lte=date_fin)


def _exporter_excel(queryset, nom_fichier):
    """Génère le classeur Excel partagé par les différents exports du journal."""
    entetes = ['Date', 'Type', 'Catégorie', 'Sous-catégorie de charge', 'Montant (FCFA)', 'Mode de paiement', 'Description']
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = 'Opérations'
    feuille.append(entetes)
    for cellule in feuille[1]:
        cellule.font = Font(bold=True, color='FFFFFF')
        cellule.fill = PatternFill('solid', fgColor='1E3260')

    for operation in queryset:
        feuille.append([
            timezone.localtime(operation.date).strftime('%d/%m/%Y %H:%M'),
            operation.get_type_operation_display(),
            operation.get_categorie_display(),
            operation.categorie_charge.nom if operation.categorie_charge else '',
            operation.montant,
            operation.mode_paiement_affiche,
            operation.description,
        ])

    for colonne in feuille.columns:
        valeurs = [str(cellule.value) for cellule in colonne if cellule.value is not None]
        largeur = max([len(v) for v in valeurs] + [10]) + 2
        feuille.column_dimensions[colonne[0].column_letter].width = largeur

    buffer = BytesIO()
    classeur.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    horodatage = timezone.now().strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}_{horodatage}.xlsx"'
    return response


class OperationBudgetListView(LoginRequiredMixin, ListView):
    """Journal des opérations budgétaires — page d'accueil du module Budget."""
    model = OperationBudget
    template_name = 'budget/operation_list.html'
    context_object_name = 'operations'
    paginate_by = 15

    def get_queryset(self):
        return _operations_filtrees(self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['type_filtre'] = self.request.GET.get('type', 'tous')
        context['categorie_filtre'] = self.request.GET.get('categorie', 'toutes')
        context['categories'] = OperationBudget.CATEGORIE_CHOICES
        context['solde_caisse'] = OperationBudget.solde_caisse()

        today = date.today()
        ops_today = OperationBudget.objects.filter(date__date=today)
        context['recettes_jour'] = sum(o.montant for o in ops_today.filter(type_operation='entree'))
        context['depenses_jour'] = sum(o.montant for o in ops_today.filter(type_operation='sortie'))

        ops_month = OperationBudget.objects.filter(date__year=today.year, date__month=today.month)
        context['recettes_mois'] = sum(o.montant for o in ops_month.filter(type_operation='entree'))
        context['depenses_mois'] = sum(o.montant for o in ops_month.filter(type_operation='sortie'))
        return context


class OperationBudgetCreateView(LoginRequiredMixin, View):
    """Saisie manuelle — uniquement charge/salaire/autre (le reste est automatique)."""
    template_name = 'budget/operation_form.html'

    def get(self, request):
        return render(request, self.template_name, {'form': OperationBudgetForm()})

    def post(self, request):
        form = OperationBudgetForm(request.POST)
        if form.is_valid():
            OperationBudget.objects.create(
                type_operation=form.cleaned_data['type_operation'],
                categorie=form.cleaned_data['categorie'],
                categorie_charge=form.cleaned_data.get('categorie_charge'),
                montant=form.cleaned_data['montant'],
                mode_paiement=form.cleaned_data['mode_paiement'],
                operateur_mobile_money=form.cleaned_data.get('operateur_mobile_money', ''),
                description=form.cleaned_data.get('description', ''),
                enregistre_par=request.user,
            )
            messages.success(request, 'Opération enregistrée avec succès.')
            return redirect('budget:operation_list')
        return render(request, self.template_name, {'form': form})


class OperationBudgetDeleteView(LoginRequiredMixin, View):
    """Suppression réservée aux opérations saisies manuellement."""

    def post(self, request, pk):
        try:
            operation = OperationBudget.objects.get(pk=pk)
        except OperationBudget.DoesNotExist:
            messages.error(request, "Opération introuvable.")
            return redirect('budget:operation_list')

        if operation.est_automatique:
            messages.error(
                request,
                "Cette opération est générée automatiquement depuis une séance, un abonnement "
                "ou un mouvement de stock — corrigez-la à la source plutôt que de la supprimer ici."
            )
        else:
            operation.delete()
            messages.success(request, 'Opération supprimée avec succès.')
        return redirect('budget:operation_list')


class OperationBudgetExportView(LoginRequiredMixin, View):
    """Export Excel du journal des opérations : le filtre actif (GET, lien) ou
    une sélection précise de lignes cochées (POST, formulaire)."""

    def get(self, request):
        queryset = _operations_filtrees(request).order_by('-date')
        return _exporter_excel(queryset, 'operations_budget')

    def post(self, request):
        ids = request.POST.getlist('ids')
        queryset = OperationBudget.objects.select_related('categorie_charge').filter(pk__in=ids).order_by('-date')
        return _exporter_excel(queryset, 'operations_budget_selection')


class RapportOperationsView(LoginRequiredMixin, ListView):
    """Rapport des opérations budgétaires sur une période choisie — vue
    synthétique (totaux entrées/sorties) distincte du journal au fil de
    l'eau (`OperationBudgetListView`)."""
    model = OperationBudget
    template_name = 'budget/rapport_operations.html'
    context_object_name = 'operations'
    paginate_by = 20

    def get_queryset(self):
        return _operations_rapport(self.request).order_by('-date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        date_debut, date_fin = _dates_rapport(self.request)
        context['date_debut'] = date_debut
        context['date_fin'] = date_fin
        context['type_filtre'] = self.request.GET.get('type', 'tous')
        context['categorie_filtre'] = self.request.GET.get('categorie', 'toutes')
        context['categories'] = OperationBudget.CATEGORIE_CHOICES

        operations_periode = _operations_rapport(self.request)
        entrees = operations_periode.filter(type_operation='entree')
        sorties = operations_periode.filter(type_operation='sortie')
        context['nb_entrees'] = entrees.count()
        context['nb_sorties'] = sorties.count()
        context['montant_entrees'] = sum(o.montant for o in entrees)
        context['montant_sorties'] = sum(o.montant for o in sorties)
        context['solde_periode'] = context['montant_entrees'] - context['montant_sorties']
        return context


class RapportOperationsExportView(LoginRequiredMixin, View):
    """Export Excel du rapport, restreint à la période et aux filtres actifs."""

    def get(self, request):
        queryset = _operations_rapport(request).order_by('-date')
        return _exporter_excel(queryset, 'rapport_operations')


def _repartition_categories(queryset):
    """Répartition par catégorie (libellé, nombre, montant) d'un ensemble
    d'opérations déjà filtré sur un sens (entrée ou sortie) — ne garde que
    les catégories réellement présentes, triées par montant décroissant."""
    repartition = []
    for valeur, libelle in OperationBudget.CATEGORIE_CHOICES:
        sous_ensemble = queryset.filter(categorie=valeur)
        nombre = sous_ensemble.count()
        if nombre:
            repartition.append({
                'libelle': libelle,
                'nombre': nombre,
                'montant': sum(o.montant for o in sous_ensemble),
            })
    repartition.sort(key=lambda ligne: ligne['montant'], reverse=True)
    return repartition


class BilanMensuelView(LoginRequiredMixin, View):
    """Bilan simplifié d'un mois choisi (activité + montants), pensé pour une
    lecture rapide sans repasser par le détail du journal — seuls les mois
    ayant au moins une opération sont proposés."""
    template_name = 'budget/bilan_mensuel.html'

    def get(self, request):
        mois_disponibles = list(OperationBudget.objects.dates('date', 'month', order='DESC'))

        mois_selectionne = None
        try:
            mois_selectionne = datetime.strptime(request.GET.get('mois', ''), '%Y-%m').date()
        except ValueError:
            pass
        if mois_selectionne not in mois_disponibles:
            mois_selectionne = mois_disponibles[0] if mois_disponibles else date.today().replace(day=1)

        operations_mois = OperationBudget.objects.filter(
            date__year=mois_selectionne.year, date__month=mois_selectionne.month,
        )
        entrees_mois = operations_mois.filter(type_operation='entree')
        sorties_mois = operations_mois.filter(type_operation='sortie')
        montant_entrees = sum(o.montant for o in entrees_mois)
        montant_sorties = sum(o.montant for o in sorties_mois)

        return render(request, self.template_name, {
            'mois_disponibles': mois_disponibles,
            'mois_selectionne': mois_selectionne,
            'nb_operations': operations_mois.count(),
            'montant_entrees': montant_entrees,
            'montant_sorties': montant_sorties,
            'solde_net': montant_entrees - montant_sorties,
            'repartition_entrees': _repartition_categories(entrees_mois),
            'repartition_sorties': _repartition_categories(sorties_mois),
        })


class CategorieChargeListView(AdminRequiredMixin, ListView):
    model = CategorieCharge
    template_name = 'budget/categorie_charge_list.html'
    context_object_name = 'categories_charge'


class CategorieChargeCreateView(AdminRequiredMixin, CreateView):
    model = CategorieCharge
    form_class = CategorieChargeForm
    template_name = 'budget/categorie_charge_form.html'
    success_url = reverse_lazy('budget:categorie_charge_list')

    def form_valid(self, form):
        messages.success(self.request, 'Catégorie de charge créée avec succès.')
        return super().form_valid(form)


class CategorieChargeUpdateView(AdminRequiredMixin, UpdateView):
    model = CategorieCharge
    form_class = CategorieChargeForm
    template_name = 'budget/categorie_charge_form.html'
    success_url = reverse_lazy('budget:categorie_charge_list')

    def form_valid(self, form):
        messages.success(self.request, 'Catégorie de charge modifiée avec succès.')
        return super().form_valid(form)


class CategorieChargeDeleteView(AdminRequiredMixin, DeleteView):
    model = CategorieCharge
    template_name = 'budget/categorie_charge_confirm_delete.html'
    success_url = reverse_lazy('budget:categorie_charge_list')

    def form_valid(self, form):
        messages.success(self.request, 'Catégorie de charge supprimée avec succès.')
        return super().form_valid(form)
